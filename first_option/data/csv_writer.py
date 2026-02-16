import openpyxl
import csv

param="address"
# param="tel"
# param="email"

excel_file = "data.xlsx"
csv_file = f"encoded_{param}_dataset.csv"
column_key = "Адрес"
# column_key = "Телефон"
# column_key ="email"

wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

header = [cell.value for cell in sheet[1]]
if column_key not in header:
    print(f"Столбец '{column_key}' не найден в Excel-файле.")
    exit()

column_index = header.index(column_key) + 1

data = [sheet.cell(row=i, column=column_index).value for i in range(2, sheet.max_row + 1)]

with open(csv_file, mode='w', newline='', encoding='utf-8') as file:
    writer = csv.writer(file)
    writer.writerow([column_key])
    for value in data:
        writer.writerow([value])

print(f"Данные из столбца '{column_key}' сохранены в CSV-файл '{csv_file}'.")
